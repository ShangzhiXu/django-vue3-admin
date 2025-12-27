from datetime import date, datetime, timedelta
from django.db.models import Q, F, ExpressionWrapper, DurationField
from django.http import HttpResponse
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from urllib.parse import quote

from dvadmin.utils.serializers import CustomModelSerializer
from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import DetailResponse
from plugins.workorder.models import WorkOrder, SupervisionPush
from dvadmin.system.utils.notifications import send_notification_to_user


class SupervisionPushWorkOrderSerializer(CustomModelSerializer):
    """
    督办中心工单列表-序列化器
    """
    merchant_name = serializers.SerializerMethodField(read_only=True)
    merchant_manager = serializers.SerializerMethodField(read_only=True)
    merchant_phone = serializers.SerializerMethodField(read_only=True)
    hazard_level_display = serializers.SerializerMethodField(read_only=True)
    inspector_name = serializers.SerializerMethodField(read_only=True)
    responsible_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_person_name = serializers.SerializerMethodField(read_only=True)
    overdue_days = serializers.SerializerMethodField(read_only=True)
    overdue_hours = serializers.SerializerMethodField(read_only=True)
    overdue_duration_display = serializers.SerializerMethodField(read_only=True)
    lag_level = serializers.SerializerMethodField(read_only=True)
    last_feedback = serializers.SerializerMethodField(read_only=True)
    
    def get_merchant_name(self, obj):
        """获取商户名称"""
        return obj.merchant.name if obj.merchant else None
    
    def get_merchant_manager(self, obj):
        """获取商户负责人"""
        return obj.merchant.manager if obj.merchant else None
    
    def get_merchant_phone(self, obj):
        """获取商户联系电话"""
        return obj.merchant.phone if obj.merchant else None
    
    def get_hazard_level_display(self, obj):
        """获取隐患等级的中文显示值"""
        return obj.get_hazard_level_display() if obj.hazard_level else None
    
    def get_inspector_name(self, obj):
        """获取检查人名称"""
        if obj.inspector:
            return obj.inspector.name
        return None
    
    def get_responsible_person_name(self, obj):
        """获取包保责任人名称"""
        if obj.responsible_person:
            return obj.responsible_person.name
        return None

    def get_transfer_person_name(self, obj):
        """获取移交负责人名称"""
        if obj.transfer_person:
            return obj.transfer_person.name
        return None
    
    def get_overdue_days(self, obj):
        """计算逾期天数"""
        if obj.deadline:
            today = date.today()
            delta = today - obj.deadline
            return max(0, delta.days)
        return 0
    
    def get_overdue_hours(self, obj):
        """计算逾期总小时数"""
        if obj.deadline:
            # 由于 USE_TZ = False，使用 naive datetime
            now = datetime.now()
            # deadline是DateField，转换为datetime（当天0点）
            deadline_datetime = datetime.combine(obj.deadline, datetime.min.time())
            if now > deadline_datetime:
                delta = now - deadline_datetime
                total_hours = int(delta.total_seconds() / 3600)
                return total_hours
        return 0
    
    def get_overdue_duration_display(self, obj):
        """获取逾期时长显示文本（如：5天4小时）"""
        days = self.get_overdue_days(obj)
        total_hours = self.get_overdue_hours(obj)
        
        if total_hours > 0:
            # 计算超出天数的额外小时数
            extra_hours = total_hours % 24
            if days > 0:
                if extra_hours > 0:
                    return f"{days}天{extra_hours}小时"
                return f"{days}天"
            else:
                return f"{total_hours}小时"
        elif days > 0:
            return f"{days}天"
        return "0小时"
    
    def get_lag_level(self, obj):
        """获取滞后级别"""
        days = self.get_overdue_days(obj)
        if days > 3:
            return {'label': '严重滞后', 'type': 'danger'}
        elif days > 1:
            return {'label': '一般滞后', 'type': 'warning'}
        elif days > 0:
            return {'label': '轻微滞后', 'type': 'info'}
        return {'label': '正常', 'type': 'success'}
    
    def get_last_feedback(self, obj):
        """获取最后反馈信息（暂时使用remark字段，后续可扩展）"""
        if obj.remark:
            return obj.remark
        return "无任何反馈"
    
    class Meta:
        model = WorkOrder
        fields = [
            "id", "workorder_no", "merchant_name", "merchant_manager", "merchant_phone",
            "hazard_level", "hazard_level_display", "problem_description",
            "report_time", "deadline", "status", "inspector_name", "responsible_person_name",
            "is_transferred", "transfer_person_name", "transfer_remark",
            "overdue_days", "overdue_hours", "overdue_duration_display",
            "lag_level", "last_feedback", "create_datetime"
        ]


class SupervisionPushSerializer(CustomModelSerializer):
    """
    督办中心记录-序列化器
    """
    workorder_count = serializers.SerializerMethodField(read_only=True)
    push_method_display = serializers.SerializerMethodField(read_only=True)
    push_status_display = serializers.SerializerMethodField(read_only=True)
    
    def get_workorder_count(self, obj):
        """获取关联工单数量"""
        return obj.workorders.count()
    
    def get_push_method_display(self, obj):
        """获取推送方式显示值"""
        return obj.get_push_method_display() if obj.push_method else None
    
    def get_push_status_display(self, obj):
        """获取推送状态显示值"""
        return obj.get_push_status_display() if obj.push_status else None
    
    class Meta:
        model = SupervisionPush
        fields = "__all__"
        read_only_fields = ["id", "push_time"]


class SupervisionWorkOrderExportSerializer(CustomModelSerializer):
    """
    督办工单导出序列化器
    """
    merchant_name = serializers.SerializerMethodField(read_only=True)
    hazard_level_display = serializers.SerializerMethodField(read_only=True)
    inspector_name = serializers.SerializerMethodField(read_only=True)
    responsible_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_person_name = serializers.SerializerMethodField(read_only=True)
    status = serializers.SerializerMethodField(read_only=True)
    overdue_duration_display = serializers.SerializerMethodField(read_only=True)
    lag_level_label = serializers.SerializerMethodField(read_only=True)
    
    def get_merchant_name(self, obj):
        """获取商户名称"""
        return obj.merchant.name if obj.merchant else ""
    
    def get_hazard_level_display(self, obj):
        """获取隐患等级的中文显示值"""
        return obj.get_hazard_level_display() if obj.hazard_level else ""
    
    def get_inspector_name(self, obj):
        """获取检查人名称"""
        if obj.inspector:
            return obj.inspector.name
        return ""
    
    def get_responsible_person_name(self, obj):
        """获取包保责任人名称"""
        if obj.responsible_person:
            return obj.responsible_person.name
        return ""
    
    def get_transfer_person_name(self, obj):
        """获取移交负责人名称"""
        if obj.transfer_person:
            return obj.transfer_person.name
        return ""
    
    def get_status(self, obj):
        """返回状态的中文显示值"""
        return obj.get_status_display() if obj.status is not None else ""
    
    def get_overdue_duration_display(self, obj):
        """获取逾期时长显示文本"""
        from datetime import date, datetime
        if obj.deadline:
            today = date.today()
            delta = today - obj.deadline
            days = max(0, delta.days)
            if days > 0:
                return f"{days}天"
        return "0小时"
    
    def get_lag_level_label(self, obj):
        """获取滞后级别标签"""
        from datetime import date
        if obj.deadline:
            today = date.today()
            delta = today - obj.deadline
            days = max(0, delta.days)
            if days > 3:
                return '严重滞后'
            elif days > 1:
                return '一般滞后'
            elif days > 0:
                return '轻微滞后'
        return '正常'
    
    class Meta:
        model = WorkOrder
        fields = (
            "workorder_no",
            "merchant_name",
            "problem_description",
            "inspector_name",
            "responsible_person_name",
            "overdue_duration_display",
            "lag_level_label",
            "hazard_level_display",
            "status",
            "deadline",
            "report_time",
        )


class SupervisionPushViewSet(CustomModelViewSet):
    """
    督办中心接口
    """
    queryset = SupervisionPush.objects.prefetch_related('workorders').all()
    serializer_class = SupervisionPushSerializer
    filter_fields = ['push_status', 'push_method']
    search_fields = ['title', 'regulatory_unit']
    
    # 督办工单列表导出配置
    export_field_label = {
        "workorder_no": "工单号",
        "merchant_name": "商户名称",
        "problem_description": "问题描述",
        "inspector_name": "检查人",
        "responsible_person_name": "包保责任人",
        "overdue_duration_display": "逾期时长",
        "lag_level_label": "滞后级别",
        "hazard_level_display": "隐患等级",
        "status": "状态",
        "deadline": "整改时限",
        "report_time": "上报时间",
    }
    export_serializer_class = SupervisionWorkOrderExportSerializer
    
    @action(methods=['get'], detail=False, url_path='workorder-list')
    def workorder_list(self, request):
        """
        获取待督办工单列表（支持筛选）
        """
        # 获取筛选参数
        overdue_hours = request.query_params.get('overdue_hours', None)  # 逾期时长（小时）
        hazard_level = request.query_params.get('hazard_level', None)  # 隐患等级
        status = request.query_params.get('status', None)  # 工单状态
        
        # 先更新逾期状态：将deadline已过但状态为待整改或待复查的工单更新为已逾期
        today = date.today()
        WorkOrder.objects.filter(
            Q(status__in=[0, 1]) &  # 待整改或待复查
            Q(deadline__lt=today) &  # 整改时限已过
            Q(deadline__isnull=False)  # deadline不为空
        ).update(status=3)
        
        # 基础查询：查询已逾期或已被督办的工单（排除已完成的工单）
        # 督办中心显示：status=3（已逾期）或 已被督办且未完成的工单
        queryset = WorkOrder.objects.exclude(status=2).filter(
            Q(status=3) |  # 已逾期状态（所有已逾期且未完成的工单）
            Q(is_supervised=True)  # 已被督办且未完成的工单（因为已经exclude了status=2，所以这里只需要is_supervised=True）
        ).select_related('merchant', 'task', 'inspector', 'responsible_person', 'transfer_person').order_by('-deadline')
        
        # 按逾期时长筛选
        if overdue_hours:
            try:
                hours = int(overdue_hours)
                # 计算最早整改时限（现在减去逾期小时数）
                # 由于 USE_TZ = False，使用 naive datetime
                cutoff_datetime = datetime.now() - timedelta(hours=hours)
                # 转换为date对象，因为deadline是DateField
                cutoff_date = cutoff_datetime.date()
                # 筛选整改时限早于cutoff_date的工单
                queryset = queryset.filter(deadline__lte=cutoff_date)
            except (ValueError, TypeError):
                pass
        
        # 按隐患等级筛选
        if hazard_level:
            queryset = queryset.filter(hazard_level=hazard_level)
        
        # 按状态筛选
        if status is not None:
            try:
                status_int = int(status)
                queryset = queryset.filter(status=status_int)
            except (ValueError, TypeError):
                pass
        
        # 自动更新逾期状态（在查询之前更新，避免重复）
        # 这部分已经在get_queryset中处理了，这里不需要重复处理
        
        # 分页
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 10))
        offset = (page - 1) * limit
        
        total = queryset.count()
        workorders = queryset[offset:offset + limit]
        
        serializer = SupervisionPushWorkOrderSerializer(workorders, many=True)
        
        return DetailResponse(
            data={
                'list': serializer.data,
                'total': total,
                'page': page,
                'limit': limit
            },
            msg="获取成功"
        )
    
    @action(methods=['post'], detail=False, url_path='batch-push')
    def batch_push(self, request):
        """
        批量推送督办通知
        """
        workorder_ids = request.data.get('workorder_ids', [])
        regulatory_unit = request.data.get('regulatory_unit', '监管单位')
        push_method = request.data.get('push_method', 'system')
        
        if not workorder_ids:
            return DetailResponse(msg="请选择要推送的工单", code=400)
        
        # 获取工单（使用 select_related 优化查询，避免 N+1 问题）
        workorders = WorkOrder.objects.filter(id__in=workorder_ids).select_related(
            'merchant', 'inspector', 'responsible_person', 'transfer_person'
        )
        if not workorders.exists():
            return DetailResponse(msg="工单不存在", code=400)
        
        # 生成推送内容
        workorder_list = []
        for wo in workorders:
            overdue_days = (date.today() - wo.deadline).days if wo.deadline else 0
            workorder_list.append(
                f"工单号：{wo.workorder_no}，商户：{wo.merchant.name if wo.merchant else '未知'}，"
                f"问题：{wo.problem_description or '无'}，逾期：{overdue_days}天"
            )
        
        push_content = f"督办通知\n\n以下是严重逾期工单，请及时处理：\n\n" + "\n".join(workorder_list)
        push_title = f"督办通知-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # 创建推送记录
        push_record = SupervisionPush.objects.create(
            title=push_title,
            regulatory_unit=regulatory_unit,
            push_method=push_method,
            push_content=push_content,
            push_status='success',  # 暂时默认为成功，后续接入实际推送接口后改为pending
            push_time=datetime.now()  # 由于 USE_TZ = False，使用 naive datetime
        )
        
        # 关联工单
        push_record.workorders.set(workorders)
        
        # 给每个工单的移交负责人、包保负责人、检查人发送小程序通知
        notification_count = 0
        
        for wo in workorders:
            overdue_days = (date.today() - wo.deadline).days if wo.deadline else 0
            merchant_name = wo.merchant.name if wo.merchant else '未知'
            
            # 通知检查人
            if getattr(wo, 'inspector', None):
                title = f"督办通知：{wo.workorder_no}"
                content = f"工单\"{wo.workorder_no}\"已被督办，请尽快处理。\n商户：{merchant_name}"
                if overdue_days > 0:
                    content += f"\n逾期：{overdue_days}天"
                send_notification_to_user(
                    user=wo.inspector,
                    title=title,
                    content=content,
                    request=request,
                    target_type=0,
                )
                notification_count += 1
            
            # 通知包保负责人
            if getattr(wo, 'responsible_person', None):
                title = f"督办通知：{wo.workorder_no}"
                content = f"工单\"{wo.workorder_no}\"已被督办，请关注整改情况。\n商户：{merchant_name}"
                if overdue_days > 0:
                    content += f"\n逾期：{overdue_days}天"
                send_notification_to_user(
                    user=wo.responsible_person,
                    title=title,
                    content=content,
                    request=request,
                    target_type=0,
                )
                notification_count += 1
            
            # 通知移交负责人（如果工单已移交）
            if getattr(wo, 'is_transferred', False) and getattr(wo, 'transfer_person', None):
                title = f"督办通知：{wo.workorder_no}"
                content = f"工单\"{wo.workorder_no}\"已被督办，请跟进处理。\n商户：{merchant_name}"
                if overdue_days > 0:
                    content += f"\n逾期：{overdue_days}天"
                send_notification_to_user(
                    user=wo.transfer_person,
                    title=title,
                    content=content,
                    request=request,
                    target_type=0,
                )
                notification_count += 1
        
        # TODO: 这里可以接入实际的推送服务（短信、邮件等）
        # 暂时只是记录，不实际发送
        
        return DetailResponse(
            data={
                'push_id': push_record.id,
                'count': workorders.count(),
                'notification_count': notification_count
            },
            msg=f"已成功推送 {workorders.count()} 个工单的督办通知，已发送 {notification_count} 条小程序通知"
        )
    
    @action(methods=['get'], detail=False, url_path='history')
    def history(self, request):
        """
        获取推送历史列表
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        # 分页
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 10))
        offset = (page - 1) * limit
        
        total = queryset.count()
        records = queryset[offset:offset + limit]
        
        serializer = self.get_serializer(records, many=True)
        
        return DetailResponse(
            data={
                'list': serializer.data,
                'total': total,
                'page': page,
                'limit': limit
            },
            msg="获取成功"
        )
    
    @action(methods=['get'], detail=False, url_path='workorder-export')
    def workorder_export(self, request):
        """
        导出督办工单列表
        """
        import datetime as dt
        
        # 获取筛选参数（与 workorder_list 保持一致）
        overdue_hours = request.query_params.get('overdue_hours', None)
        hazard_level = request.query_params.get('hazard_level', None)
        status = request.query_params.get('status', None)
        
        # 先更新逾期状态
        today = date.today()
        WorkOrder.objects.filter(
            Q(status__in=[0, 1]) &
            Q(deadline__lt=today) &
            Q(deadline__isnull=False)
        ).update(status=3)
        
        # 基础查询
        queryset = WorkOrder.objects.exclude(status=2).filter(
            Q(status=3) | Q(is_supervised=True)
        ).select_related('merchant', 'inspector', 'responsible_person', 'transfer_person').order_by('-deadline')
        
        # 应用筛选
        if overdue_hours:
            try:
                hours = int(overdue_hours)
                cutoff_datetime = datetime.now() - timedelta(hours=hours)
                cutoff_date = cutoff_datetime.date()
                queryset = queryset.filter(deadline__lte=cutoff_date)
            except (ValueError, TypeError):
                pass
        
        if hazard_level:
            queryset = queryset.filter(hazard_level=hazard_level)
        
        if status is not None:
            try:
                status_int = int(status)
                queryset = queryset.filter(status=status_int)
            except (ValueError, TypeError):
                pass
        
        # 序列化数据
        serializer = SupervisionWorkOrderExportSerializer(queryset, many=True)
        data = serializer.data
        
        # 生成Excel
        wb = Workbook()
        ws = wb.active
        
        # 表头
        headers = ["序号"] + list(self.export_field_label.values())
        ws.append(headers)
        
        # 设置表头样式
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        for idx, row_data in enumerate(data, 1):
            row = [idx] + [row_data.get(key, '') for key in self.export_field_label.keys()]
            ws.append(row)
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # 生成文件名
        timestamp = dt.datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"督办工单导出-{timestamp}.xlsx"
        
        # 创建响应
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(filename)}'
        
        wb.save(response)
        return response

