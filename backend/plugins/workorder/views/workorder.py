from datetime import date, datetime
import json
import traceback
import os
import base64
from django.db.models import Q
from django.conf import settings
from django.utils import timezone
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.renderers import JSONRenderer

from dvadmin.utils.serializers import CustomModelSerializer
from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.json_response import DetailResponse
from plugins.workorder.models import WorkOrder, WorkOrderRecheck, WorkOrderSubmission
from dvadmin.system.models import Users
from dvadmin.system.utils.notifications import send_notification_to_user


class WorkOrderSerializer(CustomModelSerializer):
    """
    工单管理-序列化器
    """
    merchant_name = serializers.SerializerMethodField(read_only=True)
    merchant_manager = serializers.SerializerMethodField(read_only=True)
    merchant_phone = serializers.SerializerMethodField(read_only=True)
    merchant_gps = serializers.SerializerMethodField(read_only=True)
    check_category_display = serializers.SerializerMethodField(read_only=True)
    hazard_level_display = serializers.SerializerMethodField(read_only=True)
    status_display = serializers.SerializerMethodField(read_only=True)
    task_name = serializers.SerializerMethodField(read_only=True)
    inspector_name = serializers.SerializerMethodField(read_only=True)
    responsible_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_person_phone = serializers.SerializerMethodField(read_only=True)
    transfer_remark = serializers.CharField(read_only=True)
    rectification_category_display = serializers.SerializerMethodField(read_only=True)
    
    def get_merchant_name(self, obj):
        """获取商户名称"""
        return obj.merchant.name if obj.merchant else None
    
    def get_merchant_manager(self, obj):
        """获取商户负责人"""
        return obj.merchant.manager if obj.merchant else None
    
    def get_merchant_phone(self, obj):
        """获取商户联系电话"""
        return obj.merchant.phone if obj.merchant else None
    
    def get_merchant_gps(self, obj):
        """获取商户GPS信息"""
        return obj.merchant.gps_status if obj.merchant and obj.merchant.gps_status else None
    
    def get_check_category_display(self, obj):
        """获取检查类别的中文显示值"""
        return obj.get_check_category_display() if obj.check_category else None
    
    def get_hazard_level_display(self, obj):
        """获取隐患等级的中文显示值"""
        return obj.get_hazard_level_display() if obj.hazard_level else None
    
    def get_status_display(self, obj):
        """获取状态的中文显示值"""
        return obj.get_status_display() if obj.status is not None else None
    
    def get_task_name(self, obj):
        """获取关联任务名称"""
        return obj.task.name if obj.task else None
    
    def get_inspector_name(self, obj):
        """获取检查人名称"""
        if obj.inspector:
            return obj.inspector.name
        return None
    
    def get_responsible_person_name(self, obj):
        """获取包保责任人名称"""
        if obj.responsible_person:
            return obj.responsible_person.name
        return None

    def get_transfer_person_name(self, obj):
        """获取移交负责人名称"""
        if obj.transfer_person:
            return obj.transfer_person.name
        return None

    def get_transfer_person_phone(self, obj):
        """获取移交负责人手机号"""
        if obj.transfer_person:
            return obj.transfer_person.mobile
        return None
    
    def get_rectification_category_display(self, obj):
        """获取整改类别的中文显示值"""
        return obj.get_rectification_category_display() if obj.rectification_category else None
    
    class Meta:
        model = WorkOrder
        fields = "__all__"
        read_only_fields = ["id", "workorder_no", "report_time"]
    
    def to_representation(self, instance):
        """序列化时自动判断是否逾期"""
        data = super().to_representation(instance)
        # 如果状态不是已逾期和已完成，且整改时限已过，自动标记为已逾期
        if instance.status not in [2, 3] and instance.deadline and instance.deadline < date.today():
            instance.status = 3
            instance.save(update_fields=['status'])
            data['status'] = 3
        return data


class WorkOrderCreateSerializer(CustomModelSerializer):
    """
    工单创建-序列化器
    """
    
    class Meta:
        model = WorkOrder
        fields = "__all__"
        read_only_fields = ["id", "workorder_no", "report_time", "status"]
    
    def create(self, validated_data):
        """创建工单时自动生成工单号，并从任务继承项目负责人"""
        # 生成工单号：WO + 年月日 + 3位序号
        today = timezone.now().date()
        date_str = today.strftime('%Y%m%d')
        
        # 查询当天最大的序号
        max_order = WorkOrder.objects.filter(
            workorder_no__startswith=f'WO{date_str}'
        ).order_by('-workorder_no').first()
        
        if max_order:
            # 提取序号并加1
            sequence = int(max_order.workorder_no[-3:]) + 1
        else:
            sequence = 1
        
        # 生成工单号
        validated_data['workorder_no'] = f'WO{date_str}{sequence:03d}'
        
        # 如果未设置检查人，且有关联任务，则从任务继承
        if not validated_data.get('inspector') and validated_data.get('task'):
            from plugins.task.models import Task
            task = validated_data['task']
            if isinstance(task, Task) and task.manager:
                validated_data['inspector'] = task.manager
        
        # 如果未设置包保责任人，且有关联商户，则从商户继承
        if not validated_data.get('responsible_person') and validated_data.get('merchant'):
            from plugins.merchant.models import Merchant
            merchant = validated_data['merchant']
            if isinstance(merchant, Merchant) and merchant.responsible_person:
                validated_data['responsible_person'] = merchant.responsible_person
        
        # 创建工单
        workorder = super().create(validated_data)

        # 创建工单后，给相关负责人发送通知
        try:
            request = self.context.get('request')
        except Exception:
            request = None

        # 通知检查人
        if getattr(workorder, 'inspector', None):
            title = f"新工单提醒：{workorder.workorder_no}"
            content = f"你被指派为工单“{workorder.workorder_no}”的检查人，请尽快处理。"
            if workorder.merchant:
                content += f"\n商户：{getattr(workorder.merchant, 'name', '')}"
            send_notification_to_user(
                user=workorder.inspector,
                title=title,
                content=content,
                request=request,
                target_type=0,
            )

        # 通知包保责任人
        if getattr(workorder, 'responsible_person', None):
            title = f"新工单提醒：{workorder.workorder_no}"
            content = f"你被指派为工单“{workorder.workorder_no}”的包保责任人，请关注整改情况。"
            if workorder.merchant:
                content += f"\n商户：{getattr(workorder.merchant, 'name', '')}"
            send_notification_to_user(
                user=workorder.responsible_person,
                title=title,
                content=content,
                request=request,
                target_type=0,
            )

        # 通知移交负责人（仅在标记为已移交且存在负责人时）
        if getattr(workorder, 'is_transferred', False) and getattr(workorder, 'transfer_person', None):
            title = f"移交工单提醒：{workorder.workorder_no}"
            content = f"工单“{workorder.workorder_no}”已移交给你，请跟进处理。"
            if workorder.merchant:
                content += f"\n商户：{getattr(workorder.merchant, 'name', '')}"
            send_notification_to_user(
                user=workorder.transfer_person,
                title=title,
                content=content,
                request=request,
                target_type=0,
            )

        return workorder


class WorkOrderUpdateSerializer(CustomModelSerializer):
    """
    工单更新-序列化器
    """
    
    class Meta:
        model = WorkOrder
        fields = "__all__"
        read_only_fields = ["id", "workorder_no", "report_time"]


class WorkOrderExportSerializer(CustomModelSerializer):
    """
    工单导出序列化器
    """
    merchant_name = serializers.SerializerMethodField(read_only=True)
    check_category_display = serializers.SerializerMethodField(read_only=True)
    inspector_name = serializers.SerializerMethodField(read_only=True)
    responsible_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_person_name = serializers.SerializerMethodField(read_only=True)
    transfer_remark = serializers.CharField(read_only=True)
    hazard_level = serializers.SerializerMethodField(read_only=True)
    status = serializers.SerializerMethodField(read_only=True)
    task_name = serializers.SerializerMethodField(read_only=True)
    rectification_category = serializers.SerializerMethodField(read_only=True)
    report_time = serializers.DateTimeField(format="%Y-%m-%d %H:%M:%S", required=False, read_only=True)
    deadline = serializers.DateField(format="%Y-%m-%d", required=False, read_only=True)
    completed_time = serializers.DateTimeField(format="%Y-%m-%d %H:%M:%S", required=False, read_only=True)
    create_datetime = serializers.DateTimeField(format="%Y-%m-%d %H:%M:%S", required=False, read_only=True)
    update_datetime = serializers.DateTimeField(format="%Y-%m-%d %H:%M:%S", required=False, read_only=True)
    # 进度信息（合并所有进度记录）
    progress_info = serializers.SerializerMethodField(read_only=True)
    
    def get_merchant_name(self, obj):
        """返回商户名称"""
        return obj.merchant.name if obj.merchant else ""
    
    def get_check_category_display(self, obj):
        """返回检查类别的中文显示值"""
        return obj.get_check_category_display() if obj.check_category else ""
    
    def get_inspector_name(self, obj):
        """返回检查人名称"""
        if obj.inspector:
            return obj.inspector.name
        return ""
    
    def get_responsible_person_name(self, obj):
        """返回包保责任人名称"""
        if obj.responsible_person:
            return obj.responsible_person.name
        return ""

    def get_transfer_person_name(self, obj):
        """返回移交负责人名称"""
        if obj.transfer_person:
            return obj.transfer_person.name
        return ""
    
    def get_hazard_level(self, obj):
        """返回隐患等级的中文显示值"""
        return obj.get_hazard_level_display() if obj.hazard_level else ""
    
    def get_status(self, obj):
        """返回状态的中文显示值"""
        return obj.get_status_display() if obj.status is not None else ""
    
    def get_task_name(self, obj):
        """返回任务名称"""
        return obj.task.name if obj.task else ""
    
    def get_rectification_category(self, obj):
        """返回整改类别的中文显示值"""
        return obj.get_rectification_category_display() if obj.rectification_category else ""
    
    def get_progress_info(self, obj):
        """返回进度信息（所有提交记录的合并文本）"""
        from plugins.workorder.models import WorkOrderSubmission
        submissions = WorkOrderSubmission.objects.filter(workorder=obj).order_by('submit_time')
        
        progress_list = []
        for sub in submissions:
            # 判断类型
            if sub.is_recheck == 1:
                type_text = "复查"
            else:
                type_text = "首次提交"
            
            # 判断是否合格
            qualified_text = "合格" if sub.is_qualified == 1 else "不合格"
            
            # 提交人
            submitter_name = sub.submitter.name if sub.submitter else "未知"
            
            # 构建进度文本
            progress_text = f"{sub.submit_time.strftime('%Y-%m-%d %H:%M:%S')} | {type_text} | {qualified_text} | 提交人：{submitter_name}"
            if sub.remark:
                progress_text += f" | 备注：{sub.remark}"
            
            progress_list.append(progress_text)
        
        # 用换行符连接所有进度
        return "\n".join(progress_list) if progress_list else ""
    
    class Meta:
        model = WorkOrder
        fields = (
            "workorder_no",
            "merchant_name",
            "check_category_display",
            "check_item",
            "inspector_name",
            "responsible_person_name",
            "is_transferred",
            "transfer_person_name",
            "transfer_remark",
            "hazard_level",
            "problem_description",
            "rectification_category",
            "report_time",
            "deadline",
            "status",
            "completed_time",
            "task_name",
            "remark",
            "progress_info",
            "create_datetime",
            "update_datetime",
        )


class WorkOrderViewSet(CustomModelViewSet):
    """
    工单管理接口
    list:查询
    create:新增
    update:修改
    retrieve:单例
    destroy:删除
    """
    queryset = WorkOrder.objects.select_related('merchant', 'task', 'inspector', 'responsible_person', 'transfer_person', 'task__manager').all()
    serializer_class = WorkOrderSerializer
    create_serializer_class = WorkOrderCreateSerializer
    update_serializer_class = WorkOrderUpdateSerializer
    filter_fields = ['status', 'hazard_level', 'deadline', 'is_transferred']
    search_fields = ['workorder_no']
    extra_filter_class = []
    
    # 导出配置
    export_field_label = {
        "workorder_no": "工单号",
        "merchant_name": "商户名称",
        "check_category_display": "检查类别",
        "check_item": "检查问题",
        "inspector_name": "检查人",
        "responsible_person_name": "包保责任人",
        "is_transferred": "是否已移交",
        "transfer_person_name": "移交负责人",
        "transfer_remark": "移交备注",
        "hazard_level": "隐患等级",
        "problem_description": "问题描述",
        "rectification_category": "整改类别",
        "report_time": "上报时间",
        "deadline": "整改时限",
        "status": "状态",
        "completed_time": "完成时间",
        "task_name": "关联任务",
        "remark": "备注",
        "progress_info": "进度记录",
        "create_datetime": "创建时间",
        "update_datetime": "更新时间",
    }
    export_serializer_class = WorkOrderExportSerializer
    
    def get_queryset(self):
        """自定义查询集，支持商户名称搜索和自动判断逾期"""
        queryset = super().get_queryset()
        
        # 自动检查并更新逾期状态（已完成状态不自动变更为逾期）
        today = date.today()
        queryset.filter(
            Q(status__in=[0, 1]) &  # 待整改或待复查
            Q(deadline__lt=today)  # 整改时限已过
        ).update(status=3)  # 更新为已逾期
        
        return queryset
    
    def filter_queryset(self, queryset):
        """自定义过滤，支持商户名称搜索"""
        queryset = super().filter_queryset(queryset)
        
        # 处理商户名称搜索
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(workorder_no__icontains=search) |
                Q(merchant__name__icontains=search)
            )
        
        # 处理上报时间范围查询
        report_time_after = self.request.query_params.get('report_time_after', None)
        report_time_before = self.request.query_params.get('report_time_before', None)
        if report_time_after:
            queryset = queryset.filter(report_time__gte=report_time_after)
        if report_time_before:
            queryset = queryset.filter(report_time__lte=report_time_before)
        
        return queryset

    @action(methods=['post'], detail=True)
    def transfer(self, request, pk=None):
        """
        设置移交负责人并标记为已移交
        """
        workorder = self.get_object()
        transfer_person_id = request.data.get('transfer_person')
        transfer_remark = request.data.get('transfer_remark', '')

        if not transfer_person_id:
            return DetailResponse(msg="请选择移交负责人", code=400)

        try:
            transfer_person = Users.objects.get(id=transfer_person_id)
        except Users.DoesNotExist:
            return DetailResponse(msg="移交负责人不存在", code=400)

        workorder.is_transferred = True
        workorder.transfer_person = transfer_person
        workorder.transfer_remark = transfer_remark
        workorder.save(update_fields=['is_transferred', 'transfer_person', 'transfer_remark', 'update_datetime'])

        # 发送通知
        send_notification_to_user(
            user=transfer_person,
            title=f"移交工单提醒：{workorder.workorder_no}",
            content=f"工单“{workorder.workorder_no}”已移交给你，请跟进处理。",
            request=request,
            target_type=0,
        )

        return DetailResponse(msg="移交成功", data={"id": workorder.id})
    
    @action(methods=['get'], detail=False, url_path='transfer-export')
    def transfer_export(self, request):
        """
        导出已移交工单列表
        """
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
        from urllib.parse import quote
        import datetime as dt
        
        # 只查询已移交的工单
        queryset = self.filter_queryset(
            self.get_queryset().filter(is_transferred=True)
        ).select_related('merchant', 'inspector', 'responsible_person', 'transfer_person', 'task')
        
        # 序列化数据
        serializer = self.export_serializer_class(queryset, many=True)
        data = serializer.data
        
        # 生成Excel
        wb = Workbook()
        ws = wb.active
        
        # 表头
        headers = ["序号"] + list(self.export_field_label.values())
        ws.append(headers)
        
        # 设置表头样式
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行
        for idx, row_data in enumerate(data, 1):
            row = [idx] + [row_data.get(key, '') for key in self.export_field_label.keys()]
            ws.append(row)
        
        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # 生成文件名
        timestamp = dt.datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"移交工单导出-{timestamp}.xlsx"
        
        # 创建响应
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Access-Control-Expose-Headers"] = "Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(filename)}'
        
        wb.save(response)
        return response
    
    @action(methods=['post'], detail=False)
    def batch_supervise(self, request):
        """
        批量督办
        """
        workorder_ids = request.data.get('ids', [])
        if not workorder_ids:
            return DetailResponse(msg="请选择要督办的工单", code=400)
        
        # 先获取工单（用于发送通知）
        workorders = WorkOrder.objects.filter(id__in=workorder_ids).select_related('responsible_person', 'merchant')
        
        # 更新状态
        count = WorkOrder.objects.filter(id__in=workorder_ids).update(
            status=0,  # 设置为待整改状态
            is_supervised=True  # 标记为已督办
        )
        
        # 只给包保责任人发送通知
        notification_count = 0
        for workorder in workorders:
            if getattr(workorder, 'responsible_person', None):
                merchant_name = workorder.merchant.name if workorder.merchant else '未知'
                overdue_days = (date.today() - workorder.deadline).days if workorder.deadline and workorder.deadline < date.today() else 0
                
                title = f"督办通知：{workorder.workorder_no}"
                content = f"工单\"{workorder.workorder_no}\"已被督办，请关注整改情况。\n商户：{merchant_name}"
                if overdue_days > 0:
                    content += f"\n逾期：{overdue_days}天"
                
                send_notification_to_user(
                    user=workorder.responsible_person,
                    title=title,
                    content=content,
                    request=request,
                    target_type=0,
                )
                notification_count += 1
        
        return DetailResponse(
            data={'count': count, 'notification_count': notification_count}, 
            msg=f"已对 {count} 个工单执行批量督办，已发送 {notification_count} 条通知"
        )
    
    @action(methods=['post'], detail=True)
    def supervise(self, request, pk=None):
        """
        单个督办
        """
        workorder = self.get_object()
        workorder.status = 0  # 设置为待整改状态
        workorder.is_supervised = True  # 标记为已督办
        workorder.save(update_fields=['status', 'is_supervised'])
        
        # 只给包保责任人发送通知
        if getattr(workorder, 'responsible_person', None):
            merchant_name = workorder.merchant.name if workorder.merchant else '未知'
            overdue_days = (date.today() - workorder.deadline).days if workorder.deadline and workorder.deadline < date.today() else 0
            
            title = f"督办通知：{workorder.workorder_no}"
            content = f"工单\"{workorder.workorder_no}\"已被督办，请关注整改情况。\n商户：{merchant_name}"
            if overdue_days > 0:
                content += f"\n逾期：{overdue_days}天"
            
            send_notification_to_user(
                user=workorder.responsible_person,
                title=title,
                content=content,
                request=request,
                target_type=0,
            )
        
        return DetailResponse(msg=f"已对工单 {workorder.workorder_no} 执行督办")
    
    @action(methods=['post'], detail=True)
    def complete(self, request, pk=None):
        """
        完成工单
        """
        workorder = self.get_object()
        # 只有非已完成状态的工单才能完成
        if workorder.status == 2:
            return DetailResponse(msg="该工单已完成", code=400)
        
        from django.utils import timezone
        workorder.status = 2  # 设置为已完成状态
        workorder.completed_time = timezone.now()  # 记录完成时间
        workorder.save(update_fields=['status', 'completed_time'])
        
        return DetailResponse(msg=f"工单 {workorder.workorder_no} 已完成")
    
    @action(methods=['get'], detail=False, url_path='transferred-list')
    def transferred_list(self, request):
        """
        获取已移交工单列表（支持筛选）
        """
        from plugins.workorder.views.supervision_push import SupervisionPushWorkOrderSerializer
        
        # 获取筛选参数
        hazard_level = request.query_params.get('hazard_level', None)  # 隐患等级
        status = request.query_params.get('status', None)  # 工单状态
        
        # 基础查询：查询已移交的工单
        queryset = WorkOrder.objects.filter(
            is_transferred=True
        ).select_related('merchant', 'task', 'inspector', 'responsible_person', 'transfer_person').order_by('-update_datetime')
        
        # 按隐患等级筛选
        if hazard_level:
            queryset = queryset.filter(hazard_level=hazard_level)
        
        # 按状态筛选
        if status is not None:
            try:
                status_int = int(status)
                queryset = queryset.filter(status=status_int)
            except (ValueError, TypeError):
                pass
        
        # 自动更新逾期状态
        today = date.today()
        queryset.filter(
            Q(status__in=[0, 1]) & Q(deadline__lt=today)
        ).update(status=3)
        
        # 分页
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 10))
        offset = (page - 1) * limit
        
        total = queryset.count()
        workorders = queryset[offset:offset + limit]
        
        serializer = SupervisionPushWorkOrderSerializer(workorders, many=True)
        
        return DetailResponse(
            data={
                'list': serializer.data,
                'total': total,
                'page': page,
                'limit': limit
            },
            msg="获取成功"
        )


class MobileWorkOrderView(APIView):
    """
    根据手机号查询负责人工单信息的API
    访问示例: /api/mobile/workorders?phone=13800138000
    返回JSON格式: {"code": 2000, "data": {...}, "msg": "查询成功"}
    """
    permission_classes = [AllowAny]  # 允许匿名访问
    renderer_classes = [JSONRenderer]  # 只返回JSON格式，不渲染HTML模板
    
    def get(self, request):
        try:
            phone = request.query_params.get('phone', None)
            
            if not phone:
                return Response({
                    "code": 400,
                    "data": [],
                    "msg": "请提供手机号参数"
                }, status=400, content_type='application/json')
            
            # 直接用手机号作为负责人索引查询工单
            # 包括：1. inspector的手机号匹配的工单 2. responsible_person的手机号匹配的工单 3. task的manager手机号匹配的工单
            workorders = WorkOrder.objects.filter(
                Q(inspector__mobile=phone) |  # 检查人手机号匹配
                Q(responsible_person__mobile=phone) |  # 包保责任人手机号匹配
                Q(transfer_person__mobile=phone) |  # 移交负责人手机号匹配
                Q(task__manager__mobile=phone)  # 任务负责人手机号匹配
            ).select_related('merchant', 'task', 'inspector', 'responsible_person', 'transfer_person', 'task__manager').order_by('-create_datetime')
            
            # 获取负责人信息（用于返回）
            manager_name = None
            manager_id = None
            # 尝试从工单的inspector、responsible_person或task的manager获取
            first_workorder = workorders.first()
            if first_workorder:
                if first_workorder.inspector and first_workorder.inspector.mobile == phone:
                    manager_name = first_workorder.inspector.name
                    manager_id = first_workorder.inspector.id
                elif first_workorder.responsible_person and first_workorder.responsible_person.mobile == phone:
                    manager_name = first_workorder.responsible_person.name
                    manager_id = first_workorder.responsible_person.id
                elif first_workorder.transfer_person and first_workorder.transfer_person.mobile == phone:
                    manager_name = first_workorder.transfer_person.name
                    manager_id = first_workorder.transfer_person.id
                elif first_workorder.task and first_workorder.task.manager and first_workorder.task.manager.mobile == phone:
                    manager_name = first_workorder.task.manager.name
                    manager_id = first_workorder.task.manager.id
            
            # 如果还是没有找到，尝试直接查询用户表
            if not manager_name:
                try:
                    user = Users.objects.filter(mobile=phone).first()
                    if user:
                        manager_name = user.name
                        manager_id = user.id
                except Exception:
                    pass
            
            # 如果没有找到任何工单和用户，返回404
            if workorders.count() == 0 and not manager_name:
                return Response({
                    "code": 404,
                    "data": [],
                    "msg": f"未找到手机号为 {phone} 的负责人工单"
                }, status=404, content_type='application/json')
            
            # 序列化工单数据
            serializer = WorkOrderSerializer(workorders, many=True)
            
            return Response({
                "code": 2000,
                "data": {
                    'phone': phone,
                    'manager_name': manager_name,
                    'manager_id': manager_id,
                    'workorders': serializer.data,
                    'total': workorders.count()
                },
                "msg": "查询成功"
            }, content_type='application/json')
        except Exception as e:
            # 捕获所有异常，确保返回JSON格式
            return Response({
                "code": 500,
                "data": [],
                "msg": f"服务器错误: {str(e)}"
            }, status=500, content_type='application/json')


class MobileWorkOrderPhotoView(APIView):
    """
    移动端工单照片上传接口
    访问示例: POST /api/mobile/workorders/{workorderId}/photos
    接收小程序发来的照片信息并输出
    """
    permission_classes = [AllowAny]  # 允许匿名访问
    renderer_classes = [JSONRenderer]  # 只返回JSON格式
    
    def post(self, request, workorder_no):
        """
        接收小程序上传的照片
        参数是工单号（如 WO20251223002）
        """
        try:
            # 只通过工单号查询工单对象
            try:
                workorder = WorkOrder.objects.get(workorder_no=workorder_no)
            except WorkOrder.DoesNotExist:
                workorder = None
            
            # 输出接收到的所有信息（无论工单是否存在都输出）
            print("=" * 80)
            if workorder:
                print(f"收到工单 {workorder_no} (工单号: {workorder.workorder_no}, ID: {workorder.id}) 的照片上传请求")
            else:
                print(f"收到工单 {workorder_no} 的照片上传请求（工单不存在）")
            print("-" * 80)
            
            # 输出请求头信息
            print("请求头信息:")
            for key, value in request.headers.items():
                print(f"  {key}: {value}")
            print("-" * 80)
            
            # 输出POST数据（非文件数据）
            print("POST数据:")
            for key, value in request.POST.items():
                print(f"  {key}: {value}")
            print("-" * 80)
            
            # 输出文件信息
            print("上传的文件:")
            uploaded_files = []
            if request.FILES:
                for key, file in request.FILES.items():
                    file_info = {
                        'field_name': key,
                        'file_name': file.name,
                        'file_size': file.size,
                        'content_type': file.content_type,
                    }
                    uploaded_files.append(file_info)
                    print(f"  字段名: {key}")
                    print(f"  文件名: {file.name}")
                    print(f"  文件大小: {file.size} bytes")
                    print(f"  内容类型: {file.content_type}")
                    print(f"  文件内容预览 (前100字节): {file.read(100)}")
                    file.seek(0)  # 重置文件指针
                    print("-" * 40)
            else:
                print("  没有上传文件")
            print("-" * 80)
            
            # 输出请求体（如果是JSON格式）
            if hasattr(request, 'data') and request.data:
                print("请求体数据 (request.data):")
                # 处理request.data，将文件对象转换为可序列化的信息
                serializable_data = {}
                for key, value in request.data.items():
                    # 检查是否是文件对象
                    if hasattr(value, 'read'):  # 文件对象
                        serializable_data[key] = {
                            'type': 'file',
                            'name': getattr(value, 'name', 'unknown'),
                            'size': getattr(value, 'size', 0),
                            'content_type': getattr(value, 'content_type', 'unknown')
                        }
                    else:
                        serializable_data[key] = value
                print(json.dumps(serializable_data, indent=2, ensure_ascii=False))
                print("-" * 80)
            
            # 输出原始请求数据
            print("原始请求数据:")
            print(f"  Content-Type: {request.content_type}")
            print(f"  Method: {request.method}")
            print(f"  Path: {request.path}")
            print(f"  Query Params: {dict(request.query_params)}")
            print("=" * 80)
            
            # 如果工单不存在，返回404
            if not workorder:
                return Response({
                    "code": 404,
                    "data": {
                        "workorder_no": workorder_no,
                        "uploaded_files": uploaded_files,
                        "post_data": dict(request.POST),
                        "message": "照片信息已接收并输出到服务器日志"
                    },
                    "msg": f"工单 {workorder_no} 不存在"
                }, status=404, content_type='application/json')
            
            # 返回成功响应
            return Response({
                "code": 2000,
                "data": {
                    "workorder_id": workorder.id,
                    "workorder_no": workorder.workorder_no,
                    "uploaded_files": uploaded_files,
                    "post_data": dict(request.POST),
                    "message": "照片信息已接收并输出到服务器日志"
                },
                "msg": "上传成功"
            }, content_type='application/json')
            
        except Exception as e:
            # 输出错误信息
            print("=" * 80)
            print(f"处理照片上传时发生错误:")
            print(f"  错误信息: {str(e)}")
            print(f"  错误详情:")
            traceback.print_exc()
            print("=" * 80)
            
            return Response({
                "code": 500,
                "data": {},
                "msg": f"服务器错误: {str(e)}"
            }, status=500, content_type='application/json')


class MobileWorkOrderCompleteView(APIView):
    """
    移动端工单完成接口
    访问示例: POST /api/mobile/workorders/{workorderNo}/complete
    通过工单号（如 WO20251223002）完成工单
    """
    permission_classes = [AllowAny]  # 允许匿名访问
    renderer_classes = [JSONRenderer]  # 只返回JSON格式
    
    def post(self, request, workorder_no):
        """
        完成工单
        参数是工单号（如 WO20251223002）
        """
        try:
            # 输出接收到的所有信息
            print("=" * 80)
            print(f"收到工单完成请求: {workorder_no}")
            print("-" * 80)
            
            # 输出请求头信息
            print("请求头信息:")
            for key, value in request.headers.items():
                print(f"  {key}: {value}")
            print("-" * 80)
            
            # 输出POST数据（非文件数据）
            print("POST数据:")
            for key, value in request.POST.items():
                print(f"  {key}: {value}")
            if not request.POST:
                print("  没有POST数据")
            print("-" * 80)
            
            # 输出请求体（如果是JSON格式）
            if hasattr(request, 'data') and request.data:
                print("请求体数据 (request.data):")
                # 处理request.data，将文件对象转换为可序列化的信息
                serializable_data = {}
                for key, value in request.data.items():
                    # 检查是否是文件对象
                    if hasattr(value, 'read'):  # 文件对象
                        serializable_data[key] = {
                            'type': 'file',
                            'name': getattr(value, 'name', 'unknown'),
                            'size': getattr(value, 'size', 0),
                            'content_type': getattr(value, 'content_type', 'unknown')
                        }
                    else:
                        serializable_data[key] = value
                print(json.dumps(serializable_data, indent=2, ensure_ascii=False))
                print("-" * 80)
            
            # 输出原始请求数据
            print("原始请求数据:")
            print(f"  Content-Type: {request.content_type}")
            print(f"  Method: {request.method}")
            print(f"  Path: {request.path}")
            print(f"  Query Params: {dict(request.query_params)}")
            print("=" * 80)
            
            # 只通过工单号查询工单对象
            try:
                workorder = WorkOrder.objects.get(workorder_no=workorder_no)
                print(f"通过工单号找到工单: {workorder.workorder_no} (ID: {workorder.id})")
            except WorkOrder.DoesNotExist:
                workorder = None
                print(f"未找到工单号: {workorder_no}")
            
            if not workorder:
                return Response({
                    "code": 404,
                    "data": {
                        "workorder_no": workorder_no,
                        "received_data": dict(request.POST) if request.POST else {},
                        "message": "工单信息已接收并输出到服务器日志"
                    },
                    "msg": f"工单 {workorder_no} 不存在"
                }, status=404, content_type='application/json')
            
            # 检查工单状态，只有非已完成状态的工单才能完成
            print(f"当前工单状态: {workorder.status} (0=待整改, 1=待复查, 2=已完成)")
            if workorder.status == 2:
                print("工单已完成，但仍会处理图片保存")
                # 即使工单已完成，也允许保存图片（可能是补充图片）
                # return Response({
                #     "code": 400,
                #     "data": {
                #         "workorder_id": workorder.id,
                #         "workorder_no": workorder.workorder_no,
                #         "status": workorder.status,
                #         "status_display": workorder.get_status_display()
                #     },
                #     "msg": "该工单已完成"
                # }, status=400, content_type='application/json')
            
            # 获取是否为复查（先判断，决定图片保存目录）
            is_recheck = request.data.get('is_recheck', 0)  # 默认为0（否）
            
            # 处理并保存图片（如果请求中包含图片）
            saved_images = []
            print("-" * 80)
            print("开始处理图片保存...")
            print(f"  is_recheck: {is_recheck}")
            print(f"  request.data 类型: {type(request.data)}")
            print(f"  request.data 内容: {list(request.data.keys()) if hasattr(request.data, 'keys') else 'N/A'}")
            
            if hasattr(request, 'data') and request.data:
                photos = request.data.get('photos', [])
                print(f"  获取到的 photos: {type(photos)}, 长度: {len(photos) if isinstance(photos, list) else 'N/A'}")
                
                if photos and isinstance(photos, list):
                    print(f"  准备保存 {len(photos)} 张图片")
                    # 根据是否为复查决定图片保存目录
                    if is_recheck == 1:
                        workorder_photo_dir = os.path.join(settings.MEDIA_ROOT, 'workorders', workorder_no, 'rechecks')
                        photo_subdir = 'rechecks'
                    else:
                        workorder_photo_dir = os.path.join(settings.MEDIA_ROOT, 'workorders', workorder_no, 'completed')
                        photo_subdir = 'completed'
                    print(f"  图片保存目录: {workorder_photo_dir}")
                    print(f"  MEDIA_ROOT: {settings.MEDIA_ROOT}")
                    
                    try:
                        os.makedirs(workorder_photo_dir, exist_ok=True)
                        print(f"  目录创建成功: {workorder_photo_dir}")
                    except Exception as e:
                        print(f"  目录创建失败: {str(e)}")
                        traceback.print_exc()
                    
                    # 处理每张图片
                    for idx, photo_data in enumerate(photos):
                        try:
                            print(f"  处理第 {idx + 1} 张图片...")
                            print(f"    图片数据类型: {type(photo_data)}")
                            print(f"    图片数据长度: {len(photo_data) if isinstance(photo_data, str) else 'N/A'}")
                            print(f"    图片数据前100字符: {photo_data[:100] if isinstance(photo_data, str) else 'N/A'}")
                            
                            # 处理base64图片数据
                            if isinstance(photo_data, str) and photo_data.startswith('data:image'):
                                print(f"    检测到base64图片数据")
                                # 提取base64数据部分
                                header, encoded = photo_data.split(',', 1)
                                print(f"    Header: {header[:50]}...")
                                print(f"    Base64数据长度: {len(encoded)}")
                                
                                # 从header中提取文件扩展名
                                if 'jpeg' in header or 'jpg' in header:
                                    ext = 'jpg'
                                elif 'png' in header:
                                    ext = 'png'
                                elif 'gif' in header:
                                    ext = 'gif'
                                else:
                                    ext = 'jpg'  # 默认使用jpg
                                print(f"    文件扩展名: {ext}")
                                
                                # 解码base64数据
                                try:
                                    image_data = base64.b64decode(encoded)
                                    print(f"    解码后数据大小: {len(image_data)} bytes")
                                except Exception as e:
                                    print(f"    Base64解码失败: {str(e)}")
                                    raise
                                
                                # 生成文件名
                                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                                if is_recheck == 1:
                                    filename = f'recheck_{idx + 1}_{timestamp}.{ext}'
                                else:
                                    filename = f'photo_{idx + 1}_{timestamp}.{ext}'
                                filepath = os.path.join(workorder_photo_dir, filename)
                                print(f"    文件路径: {filepath}")
                                
                                # 保存文件
                                try:
                                    with open(filepath, 'wb') as f:
                                        f.write(image_data)
                                    print(f"    ✓ 图片保存成功: {filepath}")
                                    
                                    # 验证文件是否存在
                                    if os.path.exists(filepath):
                                        file_size = os.path.getsize(filepath)
                                        print(f"    ✓ 文件验证成功，大小: {file_size} bytes")
                                    else:
                                        print(f"    ✗ 文件保存后不存在！")
                                except Exception as e:
                                    print(f"    ✗ 文件保存失败: {str(e)}")
                                    traceback.print_exc()
                                    raise
                                
                                # 保存相对路径（用于前端访问）
                                relative_path = f'workorders/{workorder_no}/{photo_subdir}/{filename}'
                                saved_images.append({
                                    'index': idx,
                                    'filename': filename,
                                    'url': f'/media/{relative_path}',
                                    'path': relative_path
                                })
                                print(f"    ✓ 图片信息已添加到列表")
                            else:
                                print(f"    ✗ 不是有效的base64图片数据")
                        except Exception as e:
                            print(f"  ✗ 保存图片 {idx + 1} 失败: {str(e)}")
                            traceback.print_exc()
                else:
                    print(f"  photos 为空或不是列表类型")
            else:
                print(f"  request.data 不存在或为空")
            
            print(f"  总共保存了 {len(saved_images)} 张图片")
            print("-" * 80)
            
            # 获取复查相关字段
            is_qualified = request.data.get('is_qualified', 1)  # 默认为1（合格）
            remark = request.data.get('remark', '')  # 获取备注信息
            
            print(f"  is_recheck: {is_recheck}")
            print(f"  is_qualified: {is_qualified}")
            print(f"  remark: {remark}")
            
            # 记录提交时间（无论合格与否都记录）
            submit_time = timezone.now()
            
            # 无论是否复查，都创建提交记录
            print(f"  创建提交记录（is_recheck={is_recheck}）...")
            
            # 创建提交记录
            submission_record = WorkOrderSubmission.objects.create(
                workorder=workorder,
                is_recheck=is_recheck,
                is_qualified=is_qualified,
                remark=remark,
                submit_time=submit_time
            )
            print(f"  提交记录创建成功: ID={submission_record.id}")
            
            # 根据提交结果更新工单状态（但不更新completed_time，因为completed_time只用于首次完成）
            if is_qualified == 1:
                # 合格：设置为已完成
                workorder.status = 2  # 已完成
                # 只有首次提交且合格时才更新completed_time
                if is_recheck == 0 and not workorder.completed_time:
                    workorder.completed_time = submit_time
                    print(f"  首次提交合格，设置完成时间")
                print(f"  工单状态设置为: 已完成 (status=2)")
            else:
                # 不合格：设置为待整改
                workorder.status = 0  # 待整改
                print(f"  工单状态设置为: 待整改 (status=0)")
            
            # 更新工单状态
            update_fields = ['status']
            if is_recheck == 0 and is_qualified == 1 and not workorder.completed_time:
                update_fields.append('completed_time')
            
            workorder.save(update_fields=update_fields)
            print(f"  工单更新成功")
            
            # 返回成功响应
            if is_recheck == 1:
                status_msg = "复查合格，已完成" if is_qualified == 1 else "复查不合格，需重新整改"
            else:
                status_msg = "已完成" if is_qualified == 1 else "已提交（待整改）"
            
            return Response({
                "code": 2000,
                "data": {
                    "workorder_id": workorder.id,
                    "workorder_no": workorder.workorder_no,
                    "status": workorder.status,
                    "status_display": workorder.get_status_display(),
                    "completed_time": workorder.completed_time.strftime("%Y-%m-%d %H:%M:%S") if workorder.completed_time else None,
                    "is_recheck": is_recheck,
                    "is_qualified": is_qualified,
                    "remark": remark,
                    "submission_id": submission_record.id,
                    "submit_time": submission_record.submit_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "saved_images": saved_images,
                    "images_count": len(saved_images)
                },
                "msg": f"工单 {workorder.workorder_no} {status_msg}，已保存 {len(saved_images)} 张图片"
            }, content_type='application/json')
            
        except Exception as e:
            # 输出错误信息
            print("=" * 80)
            print(f"完成工单时发生错误:")
            print(f"  错误信息: {str(e)}")
            traceback.print_exc()
            print("=" * 80)
            
            return Response({
                "code": 500,
                "data": {},
                "msg": f"服务器错误: {str(e)}"
            }, status=500, content_type='application/json')

